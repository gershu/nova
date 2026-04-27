"""
Excel report generator.

Sheets produced
---------------

1. "Top Candidates"        — best CSP across all tickers, ranked by yield.
2. "Risiko-Analyse"        — composite Risk/Reward scoring + per-ticker summary.
3. "<TICKER>"              — one sheet per watchlist entry, all surviving puts.
4. "T-Bill Matching"       — bucket -> yield mapping used in the run.
5. "Settings"              — snapshot of the settings.yaml that was active.
6. "Watchlist"             — snapshot of the watchlist that was scanned.
"""

from __future__ import annotations

import math
from collections import defaultdict
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .option_selector import CSPCandidate
from .tbill import TBillMatch, TBillMatcher
from .watchlist import Settings, WatchlistEntry


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
NUMERIC_ALIGN = Alignment(horizontal="right")
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Risk-analysis rating fills / fonts  (Excel "traffic light" palette)
_FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_FILL_ORANGE = PatternFill("solid", fgColor="FFDDC1")
_FONT_GREEN  = Font(bold=True, color="276221")
_FONT_YELLOW = Font(bold=True, color="9C5700")
_FONT_ORANGE = Font(bold=True, color="833C00")


# Column definitions: (header, attr, number_format, width)
TOP_COLUMNS = [
    ("Symbol",         "symbol",            "@",          10),
    ("Expiry",         "expiry_fmt",        "@",          12),
    ("DTE",            "dte",               "0",           7),
    ("Strike",         "strike",            "#,##0.00",   10),
    ("Spot",           "underlying_price",  "#,##0.00",   10),
    ("Moneyness",      "moneyness",         "0.00%",      11),
    ("Bid",            "bid",               "0.00",        8),
    ("Ask",            "ask",               "0.00",        8),
    ("Mid",            "mid",               "0.00",        8),
    ("Spread%",        "spread_pct",        "0.00%",      10),
    ("IV",             "iv",                "0.00%",       8),
    ("Delta",          "delta",             "0.0000",     10),
    ("Open Int.",      "open_interest",     "#,##0",      10),
    ("Volume",         "volume",            "#,##0",      10),
    ("Cash Req. (USD)", "cash_required",    "#,##0",      14),
    ("Premium (USD)",  "premium",           "#,##0.00",   13),
    ("Ann. Yield",     "annualized_yield",  "0.00%",      11),
    ("Breakeven",      "breakeven",         "#,##0.00",   11),
    ("T-Bill Bucket",  "tbill_bucket",      "0",          13),
    ("T-Bill Yield",   "tbill_yield",       "0.00%",      12),
    ("T-Bill Interest","tbill_interest",    "#,##0.00",   14),
    ("Total P/L if expires worthless", "total_if_otm", "#,##0.00", 22),
    ("Total Yield (Prem+TBill)",       "total_yield",  "0.00%",    20),
]


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def write_report(
    output_path: Path,
    candidates_by_ticker: dict[str, list[CSPCandidate]],
    tbill_matches: dict[str, dict[str, TBillMatch]],   # ticker -> expiry -> match
    matcher: TBillMatcher,
    watchlist: list[WatchlistEntry],
    settings: Settings,
    run_ts: datetime | None = None,
) -> Path:
    run_ts = run_ts or datetime.utcnow()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Default sheet -> repurpose as Top Candidates
    ws_top = wb.active
    ws_top.title = "Top Candidates"

    # ---- Aggregate top across tickers --------------------------------------
    aggregated: list[tuple[str, CSPCandidate, TBillMatch]] = []
    for ticker, candidates in candidates_by_ticker.items():
        for c in candidates:
            tb = tbill_matches.get(ticker, {}).get(c.expiry)
            aggregated.append((ticker, c, tb))
    aggregated.sort(key=lambda x: x[1].annualized_yield, reverse=True)

    _write_candidate_sheet(
        ws=ws_top,
        title=f"CSP Scanner — Top Candidates  ({run_ts.strftime('%Y-%m-%d %H:%M UTC')})",
        rows=aggregated,
    )

    # ---- Risiko-Analyse sheet -----------------------------------------------
    _write_risk_analysis_sheet(
        ws=wb.create_sheet("Risiko-Analyse"),
        aggregated=aggregated,
        tbill_matches=tbill_matches,
        run_ts=run_ts,
    )

    # ---- Per-ticker detail sheets ------------------------------------------
    for ticker, candidates in candidates_by_ticker.items():
        ws = wb.create_sheet(title=_safe_sheet_name(ticker))
        rows = [
            (ticker, c, tbill_matches.get(ticker, {}).get(c.expiry))
            for c in candidates
        ]
        _write_candidate_sheet(
            ws=ws,
            title=f"{ticker} — Cash-Secured Puts",
            rows=rows,
        )

    # ---- T-Bill matching sheet --------------------------------------------
    _write_tbill_sheet(wb.create_sheet("T-Bill Matching"), matcher, tbill_matches)

    # ---- Settings & Watchlist snapshots -----------------------------------
    _write_settings_sheet(wb.create_sheet("Settings"), settings, run_ts)
    _write_watchlist_sheet(wb.create_sheet("Watchlist"), watchlist)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _write_candidate_sheet(
    ws,
    title: str,
    rows: list[tuple[str, CSPCandidate, TBillMatch | None]],
) -> None:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TOP_COLUMNS))

    header_row = 3
    for ci, (header, _, _, width) in enumerate(TOP_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    if not rows:
        ws.cell(row=header_row + 1, column=1, value="(no candidates passed the filters)")
        return

    for ri, (ticker, cand, tbill) in enumerate(rows, start=header_row + 1):
        view = _candidate_view(cand, tbill)
        for ci, (_, attr, fmt, _) in enumerate(TOP_COLUMNS, start=1):
            value = view.get(attr)
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.number_format = fmt
            cell.border = CELL_BORDER
            if isinstance(value, (int, float)):
                cell.alignment = NUMERIC_ALIGN

    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)


def _candidate_view(c: CSPCandidate, tbill: TBillMatch | None) -> dict:
    view = asdict(c)
    view["expiry_fmt"] = _fmt_expiry(c.expiry)
    if tbill:
        view["tbill_bucket"] = tbill.bucket_days
        view["tbill_yield"] = tbill.yield_pct
        interest = tbill.interest_on(c.cash_required)
        view["tbill_interest"] = interest
        # P/L if put expires worthless: keep the premium + earn T-Bill interest
        view["total_if_otm"] = c.premium + interest
        # Combined yield on cash-secured capital, annualized
        view["total_yield"] = c.annualized_yield + tbill.yield_pct
    else:
        view["tbill_bucket"] = None
        view["tbill_yield"] = None
        view["tbill_interest"] = None
        view["total_if_otm"] = c.premium
        view["total_yield"] = c.annualized_yield
    return view


def _percentrank(values: list[float], x: float) -> float:
    """Excel-style PERCENTRANK: fraction of values strictly below x."""
    valid = [v for v in values if v is not None and not math.isnan(v)]
    if len(valid) <= 1:
        return 0.5
    below = sum(1 for v in valid if v < x)
    return below / (len(valid) - 1)


def _score_candidates(
    aggregated: list[tuple[str, "CSPCandidate", "TBillMatch | None"]],
    tbill_matches: dict,
) -> list[dict]:
    """
    Build scored rows for Risiko-Analyse.

    Score (0–100) = 50 × PERCENTRANK(ann_yield)
                  + 30 × PERCENTRANK(downside_cushion)
                  + 20 × (1 − PERCENTRANK(spread_pct))

    Rating thresholds:
      ≥ 70  →  Attraktiv   (green)
      ≥ 45  →  Mittel      (yellow)
      < 45  →  Vorsicht    (orange)
    """
    rows: list[dict] = []
    for ticker, c, tbill in aggregated:
        spot = c.underlying_price
        cushion = (
            (spot - c.strike) / spot
            if spot and not math.isnan(spot) and spot > 0
            else None
        )
        total_yield = c.annualized_yield + (tbill.yield_pct if tbill else 0.0)
        rows.append({
            "symbol":      c.symbol,
            "expiry":      _fmt_expiry(c.expiry),
            "dte":         c.dte,
            "strike":      c.strike,
            "spot":        spot if spot and not math.isnan(spot) else None,
            "cushion":     cushion,
            "breakeven":   c.breakeven,
            "ann_yield":   c.annualized_yield,
            "total_yield": total_yield,
            "iv":          c.iv if not math.isnan(c.iv) else None,
            "delta":       c.delta if not math.isnan(c.delta) else None,
            "spread":      c.spread_pct if not math.isnan(c.spread_pct) else None,
            "open_int":    c.open_interest if not math.isnan(c.open_interest) else None,
            "cash_req":    c.cash_required,
        })

    yields   = [r["ann_yield"] for r in rows if r["ann_yield"] is not None]
    cushions = [r["cushion"]   for r in rows if r["cushion"]   is not None]
    spreads  = [r["spread"]    for r in rows if r["spread"]    is not None]

    for row in rows:
        pr_y = _percentrank(yields,   row["ann_yield"]) if row["ann_yield"] is not None else 0.5
        pr_c = _percentrank(cushions, row["cushion"])   if row["cushion"]   is not None else 0.5
        pr_s = _percentrank(spreads,  row["spread"])    if row["spread"]    is not None else 0.5
        score = round(50 * pr_y + 30 * pr_c + 20 * (1 - pr_s))
        row["score"] = score
        row["rating"] = (
            "Attraktiv" if score >= 70 else
            "Mittel"    if score >= 45 else
            "Vorsicht"
        )
    return rows


_RISK_COLS = [
    # (header,            key,          number_format,  width)
    ("#",                 "#",          "0",              5),
    ("Symbol",            "symbol",     "@",             10),
    ("Expiry",            "expiry",     "@",             12),
    ("DTE",               "dte",        "0",              7),
    ("Strike",            "strike",     "#,##0.00",      10),
    ("Spot",              "spot",       "#,##0.00",      10),
    ("Downside-Puffer",   "cushion",    "0.00%",         16),
    ("Breakeven",         "breakeven",  "#,##0.00",      11),
    ("Ann. Rendite",      "ann_yield",  "0.00%",         13),
    ("Total Rendite",     "total_yield","0.00%",         14),
    ("IV",                "iv",         "0.00%",          9),
    ("Delta",             "delta",      "0.0000",        10),
    ("Spread%",           "spread",     "0.00%",         10),
    ("Open Int.",         "open_int",   "#,##0",         10),
    ("Cash Req.",         "cash_req",   "#,##0",         12),
    ("Score",             "score",      "0",              9),
    ("Bewertung",         "rating",     "@",             12),
]

_SUMMARY_COLS = [
    ("Ticker",            "@",          10),
    ("Kandidaten",        "0",          12),
    ("Ø Ann. Rendite",    "0.00%",      15),
    ("Ø Total Rendite",   "0.00%",      15),
    ("Ø Downside-Puffer", "0.00%",      17),
    ("Ø Score",           "0",          10),
    ("Bester Score",      "0",          13),
    ("Bester Breakeven",  "#,##0.00",   16),
]


def _write_risk_analysis_sheet(
    ws,
    aggregated: list[tuple[str, "CSPCandidate", "TBillMatch | None"]],
    tbill_matches: dict,
    run_ts: datetime,
) -> None:
    """
    Risiko-Analyse sheet:
      • Composite Risk/Reward score per candidate (traffic-light rating)
      • Per-ticker summary table
      • Scoring methodology note
    """
    # ---- Title ---------------------------------------------------------------
    n_cols = len(_RISK_COLS)
    ws.cell(row=1, column=1, value="Risiko-Analyse").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    subtitle = (
        f"Run: {run_ts.strftime('%Y-%m-%d %H:%M UTC')}   |   "
        "Score (0–100) = 50 % Ann. Rendite  +  30 % Downside-Puffer  +  20 % Liquidität (inv. Spread%)"
    )
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.font = Font(italic=True, size=10, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    scored = _score_candidates(aggregated, tbill_matches)
    scored.sort(key=lambda r: r["score"], reverse=True)

    if not scored:
        ws.cell(row=4, column=1, value="(Keine Kandidaten vorhanden – alle Filter haben alle Werte ausgeschlossen)")
        return

    # ---- Column headers (row 4) ---------------------------------------------
    HDR_ROW = 4
    for ci, (hdr, _, fmt, width) in enumerate(_RISK_COLS, start=1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HDR_ROW].height = 20

    # ---- Data rows -----------------------------------------------------------
    rating_fill = {"Attraktiv": _FILL_GREEN, "Mittel": _FILL_YELLOW, "Vorsicht": _FILL_ORANGE}
    rating_font = {"Attraktiv": _FONT_GREEN, "Mittel": _FONT_YELLOW, "Vorsicht": _FONT_ORANGE}

    for ri, row in enumerate(scored, start=HDR_ROW + 1):
        r_fill = rating_fill[row["rating"]]
        r_font = rating_font[row["rating"]]
        rank = ri - HDR_ROW

        for ci, (_, key, fmt, _) in enumerate(_RISK_COLS, start=1):
            value = rank if key == "#" else row.get(key)
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.number_format = fmt
            cell.border = CELL_BORDER

            if key in ("score", "rating"):
                cell.fill = r_fill
                cell.alignment = Alignment(horizontal="center")
                if key == "rating":
                    cell.font = r_font
            elif isinstance(value, (int, float)):
                cell.alignment = NUMERIC_ALIGN

    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=3)

    # ---- Per-ticker summary --------------------------------------------------
    last_data_row = HDR_ROW + len(scored)
    sum_start = last_data_row + 3

    title2 = ws.cell(row=sum_start, column=1, value="Ticker-Zusammenfassung")
    title2.font = TITLE_FONT
    ws.merge_cells(
        start_row=sum_start, start_column=1,
        end_row=sum_start, end_column=len(_SUMMARY_COLS),
    )

    HDR2 = sum_start + 2
    for ci, (hdr, fmt, width) in enumerate(_SUMMARY_COLS, start=1):
        cell = ws.cell(row=HDR2, column=ci, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    by_ticker: dict[str, list[dict]] = defaultdict(list)
    for row in scored:
        by_ticker[row["symbol"]].append(row)

    for sri, ticker in enumerate(sorted(by_ticker), start=HDR2 + 1):
        rows = by_ticker[ticker]
        ann    = [r["ann_yield"]   for r in rows if r.get("ann_yield")   is not None]
        total  = [r["total_yield"] for r in rows if r.get("total_yield") is not None]
        cush   = [r["cushion"]     for r in rows if r.get("cushion")     is not None]
        scores = [r["score"]       for r in rows if r.get("score")       is not None]
        bev    = [r["breakeven"]   for r in rows if r.get("breakeven")   is not None]

        def _avg(lst):
            return sum(lst) / len(lst) if lst else None

        values = [
            ticker,
            len(rows),
            _avg(ann),
            _avg(total),
            _avg(cush),
            round(_avg(scores)) if scores else None,
            max(scores)         if scores else None,
            max(bev)            if bev    else None,
        ]
        for ci, (val, (_, fmt, _)) in enumerate(zip(values, _SUMMARY_COLS), start=1):
            cell = ws.cell(row=sri, column=ci, value=val)
            cell.number_format = fmt
            cell.border = CELL_BORDER
            if isinstance(val, (int, float)):
                cell.alignment = NUMERIC_ALIGN

    # ---- Methodology note ---------------------------------------------------
    note_row = HDR2 + len(by_ticker) + 3
    note = ws.cell(
        row=note_row, column=1,
        value=(
            "Scoring-Methodik:  Score = 50 × PERCENTRANK(Ann. Rendite)  "
            "+  30 × PERCENTRANK(Downside-Puffer)  "
            "+  20 × (1 − PERCENTRANK(Spread%)).  "
            "Bewertung: Attraktiv ≥ 70 | Mittel ≥ 45 | Vorsicht < 45."
        ),
    )
    note.font = Font(italic=True, size=9, color="595959")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n_cols)


def _write_tbill_sheet(ws, matcher: TBillMatcher, tbill_matches: dict) -> None:
    ws.cell(row=1, column=1, value="T-Bill Matching").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    headers = ["Bucket (days)", "Yield (decimal)", "Source", "Note"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60

    for ri, (bucket, (yld, source)) in enumerate(sorted(matcher._yield_cache.items()), start=4):
        ws.cell(row=ri, column=1, value=bucket)
        c_yld = ws.cell(row=ri, column=2, value=yld)
        c_yld.number_format = "0.00%"
        ws.cell(row=ri, column=3, value=source)
        note = "Live IB quote" if source == "live" else "Fallback yield from settings.yaml"
        ws.cell(row=ri, column=4, value=note)

    # Detail of which expiry mapped to which bucket
    detail_start = ws.max_row + 3
    ws.cell(row=detail_start, column=1, value="Per-Expiry Detail").font = TITLE_FONT
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=4)
    sub_headers = ["Ticker", "Expiry", "DTE", "Bucket"]
    for ci, h in enumerate(sub_headers, start=1):
        cell = ws.cell(row=detail_start + 2, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    r = detail_start + 3
    for ticker, by_expiry in tbill_matches.items():
        for expiry, m in sorted(by_expiry.items()):
            ws.cell(row=r, column=1, value=ticker)
            ws.cell(row=r, column=2, value=_fmt_expiry(expiry))
            ws.cell(row=r, column=3, value=m.dte)
            ws.cell(row=r, column=4, value=m.bucket_days)
            r += 1


def _write_settings_sheet(ws, settings: Settings, run_ts: datetime) -> None:
    ws.cell(row=1, column=1, value="Settings Snapshot").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Run: {run_ts.strftime('%Y-%m-%d %H:%M UTC')}")

    rows: list[tuple[str, object]] = []
    rows.append(("[ib]", ""))
    for k, v in asdict(settings.ib).items():
        rows.append((f"  {k}", v))
    rows.append(("[options]", ""))
    for k, v in asdict(settings.options).items():
        rows.append((f"  {k}", v))
    rows.append(("[tbill]", ""))
    for k, v in asdict(settings.tbill).items():
        rows.append((f"  {k}", v))
    rows.append(("[report]", ""))
    for k, v in asdict(settings.report).items():
        rows.append((f"  {k}", v))

    for ri, (k, v) in enumerate(rows, start=4):
        ws.cell(row=ri, column=1, value=k)
        ws.cell(row=ri, column=2, value=str(v))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def _write_watchlist_sheet(ws, entries: list[WatchlistEntry]) -> None:
    ws.cell(row=1, column=1, value="Watchlist").font = TITLE_FONT
    headers = ["Symbol", "Exchange", "Currency", "Max Strike", "Max Contracts", "Notes"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    widths = [10, 10, 10, 14, 16, 50]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, e in enumerate(entries, start=4):
        ws.cell(row=ri, column=1, value=e.symbol)
        ws.cell(row=ri, column=2, value=e.exchange)
        ws.cell(row=ri, column=3, value=e.currency)
        ws.cell(row=ri, column=4, value=e.max_strike).number_format = "#,##0.00"
        ws.cell(row=ri, column=5, value=e.max_contracts)
        ws.cell(row=ri, column=6, value=e.notes)


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------


def _fmt_expiry(yyyymmdd: str) -> str:
    if not yyyymmdd or len(yyyymmdd) < 8:
        return yyyymmdd
    return f"{yyyymmdd[:4]}-{yyyymmdd[4:6]}-{yyyymmdd[6:8]}"


def _safe_sheet_name(name: str) -> str:
    # Excel limits: 31 chars, no [ ] : * ? / \
    bad = '[]:*?/\\'
    cleaned = "".join("_" if ch in bad else ch for ch in name)
    return cleaned[:31] or "sheet"
