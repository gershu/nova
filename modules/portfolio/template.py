"""Generates an empty Excel template for portfolio import (ConID-based).

Usage:
    python -m modules.portfolio.template /path/to/portfolio_template.xlsx

Default path (no argument):
    ~/nova_lab_input/portfolio_template.xlsx

Schema design:
  - con_id is the PRIMARY identifier (IB Contract ID, integer).
  - All instrument metadata (symbol, exchange, currency, asset_type, name)
    is auto-resolved from IB at import time.
  - User only fills:
      con_id, quantity, broker (required)
      lot_id, cost_per_share, acquired_at, account, isin, notes (optional)

Look up ConIDs in TWS:
    Trading Tools > Contract Description > Symbol search
    The Contract ID appears in the contract detail card.
"""

from __future__ import annotations

import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Excel column headers (English standard).
# Order: required fields first, then optional, then user-overrides at the end.
HEADERS = [
    "lot_id",
    "con_id",
    "isin",
    "symbol",
    "asset_class",
    "quantity",
    "cost_per_share",
    "currency",
    "acquired_at",
    "broker",
    "account",
    "name",
    "notes",
]

# Two example rows — AAPL@NASDAQ and SAP@XETRA
# (real ConIDs from IB — verified contract IDs)
EXAMPLES = [
    # lot_id   con_id   isin           sym    cls    qty  cost   ccy    acq          broker      account     name        notes
    ["LOT-001", 265598,  "US0378331005", "AAPL", "STK", 100, 175.00, "",   "2024-03-15", "IBKR",     "U1234567",  "",          ""],
    ["LOT-002", 32117,   "DE0007164600", "SAP",  "STK", 50,  130.50, "",   "2023-08-10", "comdirect", "",          "",          ""],
]

README_LINES = [
    "nova-lab Portfolio Import — Template (ConID-based)",
    "",
    "Fill in the 'Holdings' sheet, then import:",
    "    python -m modules.portfolio.import_xlsx <path-to-xlsx>",
    "",
    "Primary identifier: con_id (IB Contract ID).",
    "All instrument metadata (symbol, exchange, currency, asset_type, name)",
    "is auto-resolved from IB during import. The importer connects to IB once,",
    "queries each unique con_id, and writes the resolved metadata to the DB.",
    "",
    "Columns:",
    "  lot_id          Optional. Stable per-acquisition ID. Empty = auto-generated.",
    "  con_id          REQUIRED. Integer. IB Contract ID. Look up in TWS:",
    "                  Trading Tools > Contract Description > search by ticker.",
    "  isin            Optional. Cross-check vs IB-resolved ISIN; warning on mismatch.",
    "  symbol          Optional override. IB localSymbol. If empty, uses IB value.",
    "                  Cross-check vs IB; warning on mismatch.",
    "  asset_class     Optional override. IB-style sec-type code:",
    "                  STK / ETF / BOND / OPT / FUT / FOP / IND / CASH / CRYPTO.",
    "                  Cross-check vs IB; warning on mismatch.",
    "  quantity        REQUIRED. POST-SPLIT-ADJUSTED. Current shares held.",
    "  cost_per_share  Recommended. POST-SPLIT-ADJUSTED. Avg cost per current share.",
    "  currency        Optional override. If empty, uses IB.currency.",
    "  acquired_at     Optional. ISO date YYYY-MM-DD. Empty allowed.",
    "  broker          REQUIRED. IBKR, comdirect, DKB, Trade Republic, ...",
    "  account         Optional. Broker account ID (IB: U1234567).",
    "  name            Optional override. If empty, uses IB.longName.",
    "  notes           Optional freetext.",
    "",
    "Auto-resolved field (NOT in Excel — stored in DB after import):",
    "  exchange        IB.primaryExchange — XETRA, NASDAQ, NYSE, ...",
    "",
    "Re-import semantics:",
    "  Each import REPLACES all holdings (full sync).",
    "  Excel is single source of truth.",
    "",
    "Migration from old (symbol-based) Excel:",
    "  python -m modules.portfolio.resolve_conids <old.xlsx> <new.xlsx>",
    "  Reads old format (symbol+exchange+isin), connects IB,",
    "  produces new ConID-based xlsx.",
    "",
    "Split-adjusted values:",
    "  quantity + cost_per_share must be post-split-adjusted.",
    "  Example: 50 NVDA pre-split @ $450 after 10:1 split:",
    "    quantity = 500, cost_per_share = 45.00",
]


def write_template(target_path: pathlib.Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet 1: Holdings ---
    ws = wb.active
    ws.title = "Holdings"

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_ix, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_ix, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_ix, example in enumerate(EXAMPLES, start=2):
        for col_ix, value in enumerate(example, start=1):
            ws.cell(row=row_ix, column=col_ix, value=value)

    # Column widths heuristic
    widths = [10, 12, 16, 10, 12, 10, 16, 10, 14, 14, 12, 22, 24]
    for ix, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ix)].width = width

    ws.freeze_panes = "A2"

    # --- Sheet 2: README ---
    ws2 = wb.create_sheet("README")
    for row_ix, line in enumerate(README_LINES, start=1):
        c = ws2.cell(row=row_ix, column=1, value=line)
        if row_ix == 1:
            c.font = Font(bold=True, size=14)
        elif line.startswith("  "):
            c.font = Font(name="Menlo", size=10)
    ws2.column_dimensions["A"].width = 100

    wb.save(str(target_path))


def main() -> int:
    if len(sys.argv) > 1:
        target = pathlib.Path(sys.argv[1]).expanduser()
    else:
        target = pathlib.Path.home() / "nova_lab_input" / "portfolio_template.xlsx"

    if target.exists():
        print(f"ERROR: {target} already exists. Pass explicit path or delete file.", file=sys.stderr)
        return 64

    write_template(target)
    print(f"==> template written: {target}")
    print(f"    Fill in 'Holdings' sheet, then:")
    print(f"    python -m modules.portfolio.import_xlsx {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
