"""Migration-Helper: konvertiert altes symbol/exchange-basiertes Excel
in das neue con_id-basierte Format.

Usage:
    python -m modules.portfolio.resolve_conids <old.xlsx> <new.xlsx>

Was es macht:
  1. Liest altes Excel (akzeptiert deutsche + englische Spaltennamen via
     COLUMN_ALIASES aus import_xlsx.py).
  2. Connectet IB einmal (via IBResolver, Client-ID 12).
  3. Pro Zeile: resolve_by_symbol(symbol, exchange, currency, isin)
     → ContractDetails → ConID + IB-Metadaten.
  4. Schreibt neues Excel mit con_id-Spalte (Format wie modules.portfolio.template).
  5. Druckt Summary: rows resolved, rows failed (mit Begruendung).

Failure-Modus:
  Wenn eine Zeile nicht resolvbar ist, wird sie ins neue Excel mit leerem
  con_id geschrieben — du musst die ConID dort manuell ergaenzen.
"""

from __future__ import annotations

import pathlib
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ._ib_resolver import IBResolver, ResolvedContract
from .import_xlsx import (
    ASSET_TYPE_TO_CLASS,
    coerce_float,
    coerce_str,
    normalize_columns,
    parse_date_lenient,
)
from .template import HEADERS as NEW_HEADERS


def main() -> int:
    args = sys.argv[1:]
    if len(args) != 2:
        print("Usage: python -m modules.portfolio.resolve_conids <old.xlsx> <new.xlsx>", file=sys.stderr)
        return 64

    old_path = pathlib.Path(args[0]).expanduser().resolve()
    new_path = pathlib.Path(args[1]).expanduser().resolve()

    if not old_path.is_file():
        print(f"ERROR: {old_path} not found.", file=sys.stderr)
        return 64
    if new_path.exists():
        print(f"ERROR: {new_path} already exists — refuse to overwrite.", file=sys.stderr)
        return 64

    print("==> resolve_conids")
    print(f"    in : {old_path}")
    print(f"    out: {new_path}")

    # 1. Read old xlsx
    xl = pd.ExcelFile(old_path)
    sheet_name = "Holdings" if "Holdings" in xl.sheet_names else xl.sheet_names[0]
    df_raw = xl.parse(sheet_name)
    df = normalize_columns(df_raw)
    rows_in = len(df)
    print(f"    rows  : {rows_in}")
    print(f"    columns: {list(df.columns)}")

    # 2. Connect IB
    resolved_rows: list[dict] = []
    failures: list[tuple[int, str]] = []

    print()
    print("==> connecting IB ...")
    with IBResolver() as ib:
        print(f"    connected client_id={ib.client_id} host={ib.host} port={ib.port}")
        print()
        for ix, raw in df.iterrows():
            row = raw.to_dict()
            symbol = coerce_str(row.get("symbol"))
            exchange = coerce_str(row.get("exchange"))
            currency = coerce_str(row.get("currency"))
            isin = coerce_str(row.get("isin"), upper=True)

            line = int(ix) + 2  # Excel-Zeile (Header = 1)

            if not symbol:
                failures.append((line, "no symbol in old format"))
                resolved_rows.append(_fail_row(row, ix))
                print(f"    [SKIP] line {line}: no symbol — written with empty con_id")
                continue

            try:
                resolved: ResolvedContract | None = ib.resolve_by_symbol(
                    symbol=symbol,
                    exchange=exchange,
                    currency=currency,
                    isin=isin,
                )
            except Exception as e:  # noqa: BLE001
                failures.append((line, f"{e.__class__.__name__}: {e}"))
                resolved_rows.append(_fail_row(row, ix))
                print(f"    [FAIL] line {line} {symbol}: {e.__class__.__name__}: {e}")
                continue

            if resolved is None:
                failures.append((line, f"no IB match for {symbol}@{exchange}"))
                resolved_rows.append(_fail_row(row, ix))
                print(f"    [FAIL] line {line} {symbol}@{exchange}: no IB contract found")
                continue

            # ISIN-Cross-Check
            if isin and resolved.isin and isin.upper() != resolved.isin.upper():
                print(f"    [WARN] line {line} {symbol}: ISIN mismatch — Excel={isin} IB={resolved.isin}")

            print(f"    [OK]   line {line} {symbol:<10}@{exchange:<8} -> conId={resolved.con_id} ({resolved.name or resolved.symbol})")

            ib_class = ASSET_TYPE_TO_CLASS.get(
                resolved.asset_type or "stock",
                (resolved.asset_type or "stock").upper(),
            )
            resolved_rows.append({
                "lot_id":         coerce_str(row.get("lot_id")),
                "con_id":         resolved.con_id,
                "isin":           isin or resolved.isin or "",
                "symbol":         resolved.symbol or "",
                "asset_class":    ib_class,
                "quantity":       coerce_float(row.get("quantity")),
                "cost_per_share": coerce_float(row.get("cost_per_share")),
                "currency":       resolved.currency,
                "acquired_at":    parse_date_lenient(row.get("acquired_at")),
                "broker":         coerce_str(row.get("broker")),
                "account":        coerce_str(row.get("account")),
                "name":           coerce_str(row.get("name")) or resolved.name or "",
                "notes":          coerce_str(row.get("notes")) or "",
            })

    # 3. Write new xlsx
    _write_new_xlsx(new_path, resolved_rows)

    n_ok = rows_in - len(failures)
    print()
    print("==> done")
    print(f"    resolved : {n_ok} / {rows_in}")
    print(f"    failed   : {len(failures)}")
    if failures:
        print()
        print("    Manuelle Nacharbeit fuer die folgenden Zeilen noetig")
        print("    (im neuen Excel: con_id-Spalte ist leer, Rest erhalten):")
        for line, msg in failures[:20]:
            print(f"      line {line}: {msg}")
    print()
    print(f"    output : {new_path}")
    print(f"    Naechster Schritt:")
    print(f"      python -m modules.portfolio.import_xlsx {new_path}")
    return 0 if not failures else 2


def _fail_row(row: dict, ix: int) -> dict:
    """Wenn IB-Resolution fehlschlaegt: schreibe Zeile mit leerem con_id."""
    return {
        "lot_id":         coerce_str(row.get("lot_id")),
        "con_id":         "",
        "isin":           coerce_str(row.get("isin"), upper=True) or "",
        "symbol":         coerce_str(row.get("symbol")) or "",
        "asset_class":    "",
        "quantity":       coerce_float(row.get("quantity")),
        "cost_per_share": coerce_float(row.get("cost_per_share")),
        "currency":       coerce_str(row.get("currency"), upper=True) or "",
        "acquired_at":    parse_date_lenient(row.get("acquired_at")),
        "broker":         coerce_str(row.get("broker")) or "",
        "account":        coerce_str(row.get("account")) or "",
        "name":           coerce_str(row.get("name")) or "",
        "notes":          (coerce_str(row.get("notes")) or "") + f" [unresolved row {ix+2}]",
    }


def _write_new_xlsx(path: pathlib.Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_ix, header in enumerate(NEW_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_ix, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_ix, row in enumerate(rows, start=2):
        for col_ix, header in enumerate(NEW_HEADERS, start=1):
            v = row.get(header)
            if v is None:
                continue
            ws.cell(row=row_ix, column=col_ix, value=v)

    widths = [10, 12, 16, 10, 12, 10, 16, 10, 14, 14, 12, 22, 24]
    for ix, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ix)].width = width
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


if __name__ == "__main__":
    raise SystemExit(main())
