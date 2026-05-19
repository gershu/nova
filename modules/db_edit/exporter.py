"""DuckDB Table -> Excel xlsx.

Output-Format:
  Sheet 1: <table_name>         — Daten (Header-Row + Data-Rows)
  Sheet 2: __meta__             — Metadata fuer den Loader:
      key, value
      table_name      | <table>
      exported_at     | ISO-timestamp
      n_rows          | int
      columns         | comma-list mit type-hints "name:TYPE"
      pk_columns      | comma-list

Frozen-Row: erste Datenzeile (Header bleibt sichtbar beim Scrollen).
Spaltenbreiten: best-effort auto-fit (cap auf 60 Zeichen).
"""

from __future__ import annotations

import pathlib
from datetime import date, datetime, timezone
from typing import Optional

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema_introspect import TableInfo, get_table_info


HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
HEADER_FONT = Font(bold=True)
PK_FILL     = PatternFill("solid", fgColor="FFF2CC")    # Light yellow fuer PK-Spalten


def _format_value_for_excel(value, db_type: str):
    """Pre-format fuer xlsx-Cell. Excel ist tolerant; keep native types
    fuer numeric / date — string conversion nur bei VARCHAR."""
    if value is None:
        return None
    t = (db_type or "").upper()
    if "TIMESTAMP" in t or "DATETIME" in t:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
    if "DATE" in t:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
    return value


def export_to_xlsx(con: duckdb.DuckDBPyConnection, table: str,
                    output_path: pathlib.Path) -> tuple[int, pathlib.Path]:
    """Returns (n_rows, output_path). Raises FileNotFoundError fuer Schema oder
    duckdb.CatalogException wenn Tabelle nicht existiert."""
    info: Optional[TableInfo] = get_table_info(con, table)
    if info is None:
        raise ValueError(f"Tabelle '{table}' existiert nicht in der DB.")

    # Daten laden — sortiert nach PK-Cols wenn vorhanden, sonst keine Sort
    order_clause = ""
    if info.pk_columns:
        order_clause = "ORDER BY " + ", ".join('"' + c + '"' for c in info.pk_columns)
    col_select = ", ".join('"' + c.name + '"' for c in info.columns)
    rows = con.execute(
        f'SELECT {col_select} FROM "{table}" {order_clause}'
    ).fetchall()

    wb = Workbook()

    # ---------- Daten-Sheet ----------
    ws = wb.active
    # Sheet-Name = Tabellenname, aber Excel hat 31-Zeichen-Limit
    ws.title = table[:31]

    # Header-Row mit Type-Hint im Cell-Comment + bold + PK-Highlighting
    for col_idx, col in enumerate(info.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PK_FILL if col.is_pk else HEADER_FILL
        # Comment mit DB-Type
        from openpyxl.comments import Comment
        cell.comment = Comment(
            f"DB-Type: {col.db_type}\n"
            f"PK: {'yes' if col.is_pk else 'no'}\n"
            f"NotNull: {'yes' if col.notnull else 'no'}\n"
            + (f"Default: {col.default}\n" if col.default else ""),
            "nova-lab",
        )

    # Data-Rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (col, val) in enumerate(zip(info.columns, row), start=1):
            ws.cell(row=r_idx, column=c_idx,
                     value=_format_value_for_excel(val, col.db_type))

    # Freeze header
    ws.freeze_panes = "A2"
    # Column widths
    for c_idx, col in enumerate(info.columns, start=1):
        max_len = len(col.name)
        for r_idx in range(2, min(len(rows) + 2, 102)):     # check first 100 rows
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 60)

    # ---------- Meta-Sheet ----------
    meta_ws = wb.create_sheet("__meta__")
    meta_rows = [
        ("key",          "value"),
        ("table_name",   table),
        ("exported_at",  datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("n_rows",       str(len(rows))),
        ("columns",      ", ".join(f"{c.name}:{c.db_type}" for c in info.columns)),
        ("pk_columns",   ", ".join(info.pk_columns) if info.pk_columns else "(keine)"),
        ("notes",        "Spaltenheader = Spaltenname. PK-Spalten gelb. "
                          "Type-Hint im Header-Cell-Comment (Hover). "
                          "Loader: python -m modules.db_edit load <this-file> "
                          "[--mode truncate|insert] [--dry-run]"),
    ]
    for r_idx, (k, v) in enumerate(meta_rows, start=1):
        c_k = meta_ws.cell(row=r_idx, column=1, value=k)
        c_v = meta_ws.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            c_k.font = HEADER_FONT
            c_v.font = HEADER_FONT
            c_k.fill = HEADER_FILL
            c_v.fill = HEADER_FILL
    meta_ws.column_dimensions["A"].width = 16
    meta_ws.column_dimensions["B"].width = 120
    meta_ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(rows), output_path
