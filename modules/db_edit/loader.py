"""Excel xlsx -> DuckDB Table.

Modes:
  insert    INSERT OR REPLACE pro Zeile (append/update via PK)
  truncate  DELETE FROM <table> + INSERT, mit automatischem Backup
            der alten Rows in <table>__bkp_YYYYMMDDTHHMMSS

dry-run zeigt geplante DML ohne Commit.
Bei beliebigem Error: ROLLBACK; DB bleibt unveraendert.
"""

from __future__ import annotations

import pathlib
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Optional

import duckdb
from openpyxl import load_workbook

from .schema_introspect import (
    TableInfo,
    coerce_value,
    get_table_info,
    is_truncatable,
)


@dataclass
class LoadStats:
    table_name:        str
    mode:              str
    n_excel_rows:      int = 0
    n_db_rows_before:  int = 0
    n_inserted:        int = 0
    n_skipped:         int = 0   # rows mit allen-None
    n_db_rows_after:   int = 0
    backup_table:      Optional[str] = None
    dry_run:           bool = False
    warnings:          list[str] = field(default_factory=list)


def _read_meta(wb) -> dict:
    """Liest __meta__-Sheet (wenn vorhanden) -> dict {key: value}."""
    if "__meta__" not in wb.sheetnames:
        return {}
    ws = wb["__meta__"]
    out: dict = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            out[str(r[0]).strip()] = r[1]
    return out


def _detect_table_name(wb, override: Optional[str] = None) -> Optional[str]:
    """Reihenfolge: explizites override > __meta__-table_name > first-non-meta-sheet."""
    if override:
        return override
    meta = _read_meta(wb)
    if meta.get("table_name"):
        return str(meta["table_name"]).strip()
    for sn in wb.sheetnames:
        if sn != "__meta__":
            return sn
    return None


def _read_data_rows(wb, sheet_name: str, columns: list[str]) -> list[dict]:
    """Liest data rows; returnt list[dict {col_name: cell_value}] (raw, ohne coercion).

    Excel-Header muss exakt mit DB-Spalten-Namen matchen (case-insensitive). Extra
    Spalten im Excel werden ignoriert. Fehlende DB-Spalten -> None (DB-Default oder
    NULL).
    """
    ws = wb[sheet_name]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    excel_headers = [str(h).strip() if h else "" for h in header_cells]
    col_lower_to_idx: dict[str, int] = {}
    for idx, h in enumerate(excel_headers):
        if h:
            col_lower_to_idx[h.lower()] = idx

    out: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec: dict = {}
        all_none = True
        for col in columns:
            cell_idx = col_lower_to_idx.get(col.lower())
            v = row[cell_idx] if (cell_idx is not None and cell_idx < len(row)) else None
            rec[col] = v
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                all_none = False
        if all_none:
            continue
        rec["__row_idx__"] = row_idx
        out.append(rec)
    return out


def load_from_xlsx(con: duckdb.DuckDBPyConnection, xlsx_path: pathlib.Path,
                    *, table_override: Optional[str] = None,
                    mode: str = "insert",
                    dry_run: bool = False) -> LoadStats:
    """Mode: 'insert' (UPSERT) oder 'truncate' (DELETE+INSERT mit Backup)."""
    if mode not in ("insert", "truncate"):
        raise ValueError(f"Unsupported mode: {mode}. Use 'insert' or 'truncate'.")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        table = _detect_table_name(wb, table_override)
        if not table:
            raise ValueError(
                "Tabellenname nicht ermittelbar. Setze --table oder pflege das "
                "__meta__-Sheet mit 'table_name' im xlsx."
            )

        info = get_table_info(con, table)
        if info is None:
            raise ValueError(f"Tabelle '{table}' existiert nicht in der DB.")

        if mode == "truncate" and not is_truncatable(table):
            raise ValueError(
                f"Tabelle '{table}' ist nicht truncatable (System-/Audit-Tabelle)."
            )

        stats = LoadStats(table_name=table, mode=mode, dry_run=dry_run)
        stats.n_db_rows_before = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]

        # Welches Sheet halten wir Daten? Tabellenname (Excel-31-cap moeglich)
        data_sheet = table[:31] if table[:31] in wb.sheetnames else next(
            (s for s in wb.sheetnames if s != "__meta__"), None
        )
        if data_sheet is None:
            raise ValueError("Kein Daten-Sheet im xlsx.")

        raw_rows = _read_data_rows(wb, data_sheet, info.column_names)
        stats.n_excel_rows = len(raw_rows)

        # Coerce alle Werte gegen DB-Types — pro Zeile Liste in column-order
        coerced: list[list] = []
        coercion_errors: list[str] = []
        for rec in raw_rows:
            row_vals: list = []
            for col in info.columns:
                raw = rec.get(col.name)
                try:
                    val = coerce_value(raw, col.db_type)
                except ValueError as e:
                    coercion_errors.append(
                        f"row {rec['__row_idx__']} col '{col.name}': {e}"
                    )
                    val = None
                # NotNull-Validation (warning, nicht abort)
                if col.notnull and val is None and col.default is None:
                    stats.warnings.append(
                        f"row {rec['__row_idx__']}: NOT-NULL-col '{col.name}' "
                        f"hat keinen Wert (DB wird Fehler werfen)"
                    )
                row_vals.append(val)
            coerced.append(row_vals)

        if coercion_errors:
            print("\nCoercion-Errors:")
            for e in coercion_errors[:20]:
                print(f"  {e}")
            if len(coercion_errors) > 20:
                print(f"  ... ({len(coercion_errors) - 20} more)")
            raise ValueError(f"{len(coercion_errors)} Coercion-Errors; ABORT.")

        # Backup before truncate
        if mode == "truncate" and not dry_run and stats.n_db_rows_before > 0:
            ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            backup_name = f"{table}__bkp_{ts}"
            con.execute(f'CREATE TABLE "{backup_name}" AS SELECT * FROM "{table}"')
            stats.backup_table = backup_name

        if dry_run:
            print(f"\n[DRY-RUN] Mode '{mode}': wuerde {stats.n_excel_rows} Excel-Rows "
                  f"applizieren auf '{table}' (current: {stats.n_db_rows_before} rows).")
            if mode == "truncate" and stats.n_db_rows_before > 0:
                print(f"[DRY-RUN] Backup nach '{table}__bkp_<ts>' wuerde {stats.n_db_rows_before} rows snapshotten.")
            stats.n_db_rows_after = stats.n_db_rows_before
            return stats

        # ---------- Apply ----------
        con.begin()
        try:
            if mode == "truncate":
                con.execute(f'DELETE FROM "{table}"')

            placeholders = ", ".join(["?"] * len(info.columns))
            col_list = ", ".join(f'"{c.name}"' for c in info.columns)
            sql = f'INSERT OR REPLACE INTO "{table}" ({col_list}) VALUES ({placeholders})'
            for vals in coerced:
                con.execute(sql, vals)
                stats.n_inserted += 1
            con.commit()
        except Exception:
            con.rollback()
            raise

        stats.n_db_rows_after = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
        return stats
    finally:
        wb.close()
